import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def format_excel_from_df(df: pd.DataFrame, output_path: str):
    wb = Workbook()
    ws = wb.active

    # --- Adding columns ---
    ws.append(list(df.columns))

    # --- Adding data from columns ---
    for row in df.itertuples(index=False):
        ws.append(row)

    # Table size
    max_row = ws.max_row
    max_col = ws.max_column

    # --- Styling ---
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # --- First line editing ---
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = gray_fill
        cell.font = bold_font
        cell.alignment = center_alignment

    # --- Adding borders ---
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell.border = thin_border

    # --- Setting up column width ---
    for col in range(1, max_col + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in range(1, max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 2

    # --- First line freezing ---
    ws.freeze_panes = "A2"

    # File saving
    wb.save(output_path)
    print(f"Файл сохранен как: {output_path}")



